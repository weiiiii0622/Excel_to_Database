from openpyxl import load_workbook
import os, django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "work.settings")
django.setup()

from excel.models import User, Info, Lesson

def import_xl():
    wb = load_workbook('test.xlsx')
    ws = wb.get_sheet_by_name('test')

    max_column = ws.max_column
    max_row = ws.max_row

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_column):
        name, school, grade, gender = [row[i].value for i in range(0,4)]
        sport, food, housework, design, med, art, business, other = [row[i].value for i in range(4,max_column-1)]

        user = User.objects.get_or_create(
            name = str(name),
            school = str(school),
            grade = str(grade),
            gender = str(gender)
        )
        
        for i in range(4,max_column-1):
            try:
                for lesson in row[i].value.split(', '):
                    print(str(lesson))
            except:
                pass




if __name__ == "__main__":
    import_xl()